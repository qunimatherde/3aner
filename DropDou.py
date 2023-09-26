import pandas as pd
from openpyxl import load_workbook

# 加载Excel工作簿
workbook = load_workbook('August1.xlsx')

# 遍历每个工作表
for sheet_name in workbook.sheetnames:
    # 加载工作表数据到pandas DataFrame
    sheet = workbook[sheet_name]
    data = sheet.values
    cols = next(data, None)  # 当迭代器为空时使用None作为默认值
    if cols is None:
        continue  # 如果工作表中没有行，则跳过到下一个工作表

    df = pd.DataFrame(data, columns=cols)

    # 删除重复的行
    df.drop_duplicates(inplace=True)

    # 清空工作表内容
    sheet.delete_rows(1, sheet.max_row)

    # 将处理后的DataFrame数据写入工作表
    for index, row in df.iterrows():
        sheet.append(row.tolist())
# 保存工作簿
workbook.save('your_workbook.xlsx')
